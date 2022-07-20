
from mordecai import Geoparser
geo=Geoparser()

from openpyxl import Workbook
from openpyxl import load_workbook


def generateMordecaiLocations(worksheet):

    file = open("MordecaiLocationsLower.txt","w",encoding="utf-8")
    
    #all spreadsheets start at row 5
    start=5
    end=worksheet.max_row

    while start<end:
        #first we extract the location from the line
        text=worksheet['F'+str(start)].value
        try:
           #infofound will be a dictionary containing all the relevant info geoparse found in the string
            infofound=geo.geoparse(text)
        except:
            infofound=False
        if text is None:
            text="NOTEXTFOUND"
        file.write("____________________ \n"+
                "Row: "+str(start)+"\n"+
                "Card #: "+str(worksheet['AC'+str(start)].value)+"\n"
                "Original Text: "+text+"\n")
        if infofound:
            for entry in infofound:
                file.write("\n")
                keys=entry.keys()
                for key in keys:
                    data= entry[key]
                    if type(data) is dict:
                        datakeys=data.keys()
                        for datakey in datakeys:
                            file.write(datakey+" : "+str(data[datakey])+"\n")
                    else:
                        file.write(key+" : "+str(entry[key])+"\n")
                    
        start=start+1              

    file.close()
    return


workbook=load_workbook('Card_Upload_Template_Output.xlsx')
worksheet=workbook['Data Fields']

generateMordecaiLocations(worksheet)