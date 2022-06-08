#This will NOT work on teton at the moment because it does not have openpyxl
import re
import os
#this got auto added? idk what it is
from xml.etree.ElementTree import tostring
#openpyxl should let us edit and access excel files
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('CARD Upload Template.xlsx')
ws =wb['Data Fields']

#this way so we can iterate over fields
fields={"Location: ":"F", "Material Dated: ":"C",
"Lab Name: ":False,"Lab Number ":"A","Age: ":"Q","Age Sigma: ":"R","Latitiude ":"G","Longitude ":"H","Type Of Date: ":"E"}


#These are all of the
#  possible fields
location = ""
materialDated = ""
labName = ""
labNumber = ""
age = ""
ageSigma = ""
latitude = ""
longitude = ""
typeOfDate = ""
siteIdentifier = ""

ageAssigned = 0


#Where our data will come from and go to
input = 'D:\\ARCC\\radiocarbon\\radiocarbonOutputToSpreadsheet\\input'
output = 'D:\\ARCC\\radiocarbon\\radiocarbonOutputToSpreadsheet\\output'

#counters to know what row and column we are at
#col='A'
row=5

for file in os.listdir(input):
    print(file)
    
    #obtain the file number and store it in filenum, then put that in our spreadsheet
    filenum =file.split("_")[0]
    ws["AC"+str(row)]=str(filenum)

    #open the file and iterate over its lines
    with open(os.path.join(input,file), "r") as text_file:
        for line in text_file:
            #use this for looping. doing it this way so i can easily access index of strings in fields
            
            for key in fields:
                #check if each field is found on this line
                field_found=re.search(key, line)
                if field_found:
                    #make sure the key has an entry
                    if fields[key]:
                        #remove the title and then the first space from the line
                        content=line.split(":")[1]
                        content=content[1:]
                        #65=A
                        col=fields[key]
                        ws[col+str(row)]=str(content)

                    #we can also break once the field is found since there is only one per line
                    break
                
    #using ascii we can increment columns
    #col=chr(ord(col) + 3)

    #increment to the next entry for the next card
    row=row+1


#finally we save the spreadsheet
wb.save('Card Upload Template Output.xlsx')
